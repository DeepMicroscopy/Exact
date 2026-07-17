import csv
import io
import json
import re as _re

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST

from exact.images.models import ImageSet, Image

from .models import TableDataset, TableVersion, TableViewSettings


# ---------------------------------------------------------------------------
# Permission helper
# ---------------------------------------------------------------------------

def _can_read(imageset, user):
    return imageset.has_perm('read', user)


def _can_write(imageset, user):
    return imageset.has_perm('annotate', user)


# ---------------------------------------------------------------------------
# List / create
# ---------------------------------------------------------------------------

@login_required
def table_list(request, imageset_id):
    imageset = get_object_or_404(ImageSet, pk=imageset_id)
    if not _can_read(imageset, request.user):
        return HttpResponse('Forbidden', status=403)

    if request.method == 'POST':
        if not _can_write(imageset, request.user):
            return HttpResponse('Forbidden', status=403)
        name = request.POST.get('name', '').strip()
        if not name:
            return HttpResponse('Name required', status=400)
        dataset = TableDataset.objects.create(
            image_set=imageset,
            name=name,
            description=request.POST.get('description', '').strip(),
            creator=request.user,
        )
        return redirect(reverse('tables:view', args=[dataset.pk]))

    datasets = (TableDataset.objects
                .filter(image_set=imageset)
                .prefetch_related('versions'))
    return render(request, 'tables/list.html', {
        'imageset': imageset,
        'datasets': datasets,
    })


# ---------------------------------------------------------------------------
# Editor page
# ---------------------------------------------------------------------------

@login_required
def table_view(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return HttpResponse('Forbidden', status=403)

    current = dataset.current_version()
    return render(request, 'tables/view.html', {
        'dataset': dataset,
        'imageset': dataset.image_set,
        'current_version': current,
        'can_write': _can_write(dataset.image_set, request.user),
    })


# ---------------------------------------------------------------------------
# Data API
# ---------------------------------------------------------------------------

@login_required
@require_GET
def table_data_api(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    current = dataset.current_version()
    return JsonResponse({
        'csv': dataset.get_current_csv(),
        'version': current.version_number if current else 0,
    })


@login_required
@require_POST
def table_save_api(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_write(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    csv_text = body.get('csv', '')
    comment = body.get('comment', '').strip()

    if not csv_text.strip():
        return JsonResponse({'error': 'Empty data'}, status=400)

    version = dataset.save_new_version(csv_text, user=request.user, comment=comment)
    return JsonResponse({
        'ok': True,
        'version': version.version_number,
        'row_count': version.row_count,
        'col_count': version.col_count,
    })


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

@login_required
@require_POST
def table_import_api(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_write(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    try:
        csv_text = uploaded.read().decode('utf-8-sig')  # handle BOM
    except UnicodeDecodeError:
        try:
            uploaded.seek(0)
            csv_text = uploaded.read().decode('latin-1')
        except Exception:
            return JsonResponse({'error': 'Cannot decode file'}, status=400)

    # Determine delimiter (default comma; accept \t for tab)
    raw_delim = request.POST.get('delimiter', ',')
    if raw_delim == '\\t':
        raw_delim = '\t'
    delimiter = raw_delim if len(raw_delim) == 1 else ','

    # Parse with requested delimiter
    try:
        reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
        rows = list(reader)
    except Exception as e:
        return JsonResponse({'error': f'CSV parse error: {e}'}, status=400)

    if len(rows) < 1:
        return JsonResponse({'error': 'CSV is empty'}, status=400)

    # If data has no header row, prepend generated column names
    has_header = request.POST.get('has_header', '1') != '0'
    if not has_header:
        col_count = max((len(r) for r in rows), default=0)
        generated = [f'Col {i + 1}' for i in range(col_count)]
        rows = [generated] + rows

    # Re-serialise as standard comma-delimited CSV for storage
    if delimiter != ',' or not has_header:
        buf = io.StringIO()
        csv.writer(buf).writerows(rows)
        csv_text = buf.getvalue()

    comment = request.POST.get('comment', 'CSV import').strip() or 'CSV import'
    has_existing = dataset.current_version() is not None
    force = request.POST.get('force', 'false').lower() == 'true'

    if has_existing and not force:
        return JsonResponse({
            'conflict': True,
            'message': 'Table already has data. Confirm to create a new version.',
            'row_count': len(rows) - 1,
            'col_count': len(rows[0]) if rows else 0,
        })

    version = dataset.save_new_version(csv_text, user=request.user, comment=comment)
    return JsonResponse({
        'ok': True,
        'version': version.version_number,
        'row_count': version.row_count,
        'col_count': version.col_count,
    })


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@login_required
@require_GET
def table_export_csv(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return HttpResponse('Forbidden', status=403)

    csv_text = dataset.get_current_csv()
    response = HttpResponse(csv_text, content_type='text/csv; charset=utf-8')
    filename = dataset.name.replace(' ', '_') + '.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_GET
def table_export_xlsx(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return HttpResponse('Forbidden', status=403)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    csv_text = dataset.get_current_csv()
    rows = list(csv.reader(io.StringIO(csv_text)))

    # Load per-user column widths
    settings_obj = TableViewSettings.objects.filter(
        dataset=dataset, user=request.user).first()
    col_widths = {}
    if settings_obj:
        col_widths = settings_obj.settings_json.get('col_widths', {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = dataset.name[:31]

    header_fill = PatternFill('solid', fgColor='2D0A3E')
    header_font = Font(bold=True, color='FFFFFF')
    link_font   = Font(color='8B5CF6', underline='single')

    _EXACT_REF_RE = _re.compile(
        r'^(?:https?://[^/]+)?(/(?:images/imageset|annotations)/\d+/)$'
    )
    base_url = request.build_absolute_uri('/').rstrip('/')

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            display_val = val
            hyperlink_url = None
            if r_idx > 1 and val:
                m = _EXACT_REF_RE.match(val)
                if m:
                    hyperlink_url = base_url + m.group(1)
                    display_val = hyperlink_url
            cell = ws.cell(row=r_idx, column=c_idx, value=display_val)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            elif hyperlink_url:
                cell.hyperlink = hyperlink_url
                cell.font = link_font
            # Apply column width (px → Excel char units ≈ px/7)
            px = col_widths.get(str(c_idx - 1))
            if px:
                ws.column_dimensions[get_column_letter(c_idx)].width = max(8, int(px) // 7)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = dataset.name.replace(' ', '_') + '.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Version history
# ---------------------------------------------------------------------------

@login_required
@require_GET
def table_versions_api(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    versions = dataset.versions.select_related('created_by').order_by('-version_number')
    data = []
    for v in versions:
        data.append({
            'version': v.version_number,
            'created_at': v.created_at.strftime('%Y-%m-%d %H:%M'),
            'created_by': (v.created_by.get_full_name().strip()
                           or v.created_by.username) if v.created_by else '—',
            'comment': v.comment,
            'row_count': v.row_count,
            'col_count': v.col_count,
            'is_current': v.is_current,
        })
    return JsonResponse({'versions': data})


@login_required
@require_GET
def table_version_data_api(request, dataset_id, version_number):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    try:
        csv_text = dataset.get_version_csv(version_number)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'csv': csv_text, 'version': version_number})


# ---------------------------------------------------------------------------
# Per-user view settings
# ---------------------------------------------------------------------------

@login_required
def table_settings_api(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_read(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    if request.method == 'GET':
        obj = TableViewSettings.objects.filter(
            dataset=dataset, user=request.user).first()
        return JsonResponse(obj.settings_json if obj else {})

    if request.method == 'PUT':
        try:
            payload = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        obj, _ = TableViewSettings.objects.get_or_create(
            dataset=dataset, user=request.user)
        obj.settings_json = payload
        obj.save()
        return JsonResponse({'ok': True})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ---------------------------------------------------------------------------
# Delete dataset
# ---------------------------------------------------------------------------

@login_required
@require_GET
def resolve_url_api(request):
    """Resolve an internal EXACT path to a human-readable name and type."""
    path = request.GET.get('path', '').strip().rstrip('/')

    m = _re.match(r'^/?images/imageset/(\d+)$', path)
    if m:
        imageset = ImageSet.objects.filter(pk=int(m.group(1))).first()
        if not imageset or not imageset.has_perm('read', request.user):
            return JsonResponse({'error': 'Not found'}, status=404)
        return JsonResponse({'type': 'imageset', 'name': imageset.name,
                             'detail': imageset.team.name if imageset.team else ''})

    m = _re.match(r'^/?annotations/(\d+)$', path)
    if m:
        image = (Image.objects
                 .select_related('image_set', 'image_set__team')
                 .filter(pk=int(m.group(1))).first())
        if not image or not image.image_set.has_perm('read', request.user):
            return JsonResponse({'error': 'Not found'}, status=404)
        return JsonResponse({'type': 'image', 'name': image.name,
                             'detail': image.image_set.name})

    return JsonResponse({'error': 'Not an EXACT reference'}, status=400)


@login_required
@require_GET
def ref_picker_imagesets_api(request):
    """Return all imagesets accessible to the user for the reference picker."""
    imagesets = (ImageSet.objects
                 .filter(team__in=request.user.team_set.all())
                 .select_related('team')
                 .order_by('team__name', 'name')
                 .values('id', 'name', 'team__name')[:300])
    return JsonResponse({'results': [
        {'id': s['id'], 'name': s['name'], 'team': s['team__name']} for s in imagesets
    ]})


@login_required
@require_GET
def ref_picker_images_api(request, imageset_id):
    """Return images inside an imageset for the reference picker."""
    imageset = get_object_or_404(ImageSet, pk=imageset_id)
    if not imageset.has_perm('read', request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    images = list(imageset.images.order_by('name').values('id', 'name')[:500])
    return JsonResponse({'results': images})


@login_required
@require_POST
def table_delete_api(request, dataset_id):
    dataset = get_object_or_404(TableDataset, pk=dataset_id)
    if not _can_write(dataset.image_set, request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    imageset_id = dataset.image_set_id
    dataset.delete()
    return JsonResponse({'ok': True, 'redirect': reverse('tables:list', args=[imageset_id])})
